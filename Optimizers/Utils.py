import time
import pandas as pd
import pprint
import os
import sys
import itertools


def hamming_distance(i1, i2, numNodes):
    if len(i1) != len(i2):
        return 0

    z1, z2 = [0] * numNodes,  [0] * numNodes
    distance = 0

    # khoi tao z1, z2
    for i in range (0, len(i1)-1):
        if i1[i] <= numNodes and i1[i+1] <= numNodes:
            z1[i1[i]-1] = i1[i+1]
        elif i1[i] <= numNodes or i1[i+1] <= numNodes:
            index = min(i1[i], i1[i+1])
            z1[index-1] = -1
        if i == len(i1)-2 and i1[i+1] <= numNodes:
            z1[i1[i+1]-1] = -1

    for i in range (0, len(i2)-1):
        if i2[i] <= numNodes and i2[i+1] <= numNodes:
            z2[i2[i]-1] = i2[i+1]
        elif i2[i] <= numNodes or i2[i+1] <= numNodes:
            index = min(i2[i], i2[i+1])
            z2[index-1] = -1
        if i == len(i2)-2 and i2[i+1] <= numNodes:
            z2[i2[i+1]-1] = -1

    for i in range(numNodes):
        if z1[i] != z2[i]:
            distance += 1

    return distance

def to_excel(filename, data_ws, df, start_row=2, start_col=2):
    """Replacement for pandas .to_excel(). 
    For .xlsx and .xls formats only.
    args:
        start_row: df row +2; does not include header and is 1 based indexed.
    """
    writer = pd.ExcelWriter(filename.lower(), engine='openpyxl')
    import openpyxl
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    if data_ws not in wb.sheetnames:
        wb.create_sheet(data_ws)

    # Create the worksheet if it does not yet exist.
    writer.book = wb
    writer.sheets = {x.title: x for x in wb.worksheets}

    ws = writer.sheets[data_ws]
    # Fill with blanks.
    try:
        for row in ws:
            for cell in row:
                cell.value = None
    except TypeError:
        pass

    # Write manually to avoid overwriting formats.

    # Column names.
    ws.cell(1, 1).value = df.columns.name
    for icol, col_name in zip(range(2, len(df.columns) + 2), df.columns):
        ws.cell(1, icol).value = col_name

    # Row headers.
    for irow, row_name in zip(range(2, len(df.index) + 2), df.index):
        ws.cell(irow, 1).value = row_name

    # Body cells.
    for row, irow in zip([x[1] for x in df.iloc[start_row - 2:].iterrows()], list(range(start_row, len(df.index) + 2))):
        for cell, icol in zip([x[1] for x in row.iloc[start_col - 2:].items()], list(range(start_col, len(df.columns) + 2))):
            ws.cell(irow, icol).value = cell  # Skip the index.

    for row in ws.values:
        print('\t'.join(str(x or '') for x in row))
    print('Saving.')
    while True:
        try:
            writer.save()
            break
        except PermissionError:
            print(f'Please close {filename} before we can write to it!')
            time.sleep(2)
    writer.close()
    print('Done saving df.')

def save_solution(instance, run_time, number_of_tech, cost, work_time, version, params, t_route, uav_route, current_time = ''):
    # Instance, run_time, number_of_tech, cost, work_time
    output_file = 'result.xlsx'
    if current_time != '':
        output_file = f'./result3/new/{current_time}result.xlsx'

    data=[]
    try:
        df = pd.read_excel(output_file, sheet_name='result', engine='openpyxl')
        
        for i in range(len(df['Instance'])):
            data.append([
                df['Instance'][i], 
                df['run_time'][i], 
                df['number_of_tech'][i], 
                df['cost'][i], 
                df['work_time'][i], 
                df['version'][i], 
                df['params'][i],
                df['t_route'][i], 
                df['uav_route'][i]
            ])
    except FileNotFoundError:
        print(f"No such file or directory: {output_file}")
    data.append([instance, run_time, number_of_tech, cost, work_time, version, params, t_route, uav_route])
    df = pd.DataFrame(data, columns=['Instance', 'run_time', 'number_of_tech', 'cost', 'work_time', 'version', 'params', 't_route', 'uav_route'])
    to_excel(output_file, 'result', df)

def save_stats(instance, version, run_time, tech_num, work_time, level, record, params, current_time=''):

    # Instance, tech_num, work_time, iter0, iter1, ...
    if not (level == 'upper' or level == 'lower' or level == 'diff'):
        print("param lever must be either 'upper' or 'lower'")
        return

    output_file = 'stats.xlsx'
    if current_time != '':
        output_file = f'./result3/new/{current_time}stats.xlsx'

    data=[]
    new_data=None
    col_num = 0
    iter_cols = None

    if not os.path.exists(output_file):
        df = pd.DataFrame()
        writer = pd.ExcelWriter(output_file, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='upper', index=False)
        df.to_excel(writer, sheet_name='lower', index=False)
        df.to_excel(writer, sheet_name='diff', index=False)
        writer.save()
    
    try:
        df = pd.read_excel(output_file, sheet_name=level, engine='openpyxl')
        old_data = df.values.tolist()
        data.extend(list(map(lambda i: old_data[i][1:], range(len(old_data)))))
        col_num = len(record) + 6 if len(record) + 6 > df.columns.shape[0]-1 else df.columns.shape[0]-1
    except FileNotFoundError:
        print(f"No such file or directory: {output_file}")
    data.append([instance, version, run_time, tech_num, work_time, params, *record])
    iter_cols = list(map(lambda x: f'iter{x}', range(col_num-6)))
    empty_data = [['-']*col_num]*len(data)
    new_data = list(map(lambda i: data[i] + empty_data[i][len(data[i]):len(empty_data[i])], range(len(empty_data))))
    
    df1 = pd.DataFrame(new_data, columns=['Instance', 'version', 'run_time', 'tech_num', 'work_time', 'params', *iter_cols])
    to_excel(output_file, level, df1)

def adjust_results(filename, output_file):
    data = []
    try:
        df = pd.read_excel(filename, sheet_name='result', engine='openpyxl')
        for i in range(len(df['Instance'])):
            data.append([
                df['Instance'][i], 
                df['run_time'][i], 
                df['cost'][i], 
                df['work_time'][i], 
                str(df['version'][i])+'|'+str(df['params'][i])+'|'+ str(df['number_of_tech'][i])
            ])
        df1= pd.DataFrame(data, columns=['Instance', 'run_time', 'cost', 'work_time', 'params'])
        to_excel(output_file, 'result', df1)
    except FileNotFoundError:
        print(f"No such file or directory: {output_file}")

def combine_results(folder_path, output_path):
    current_time = time.strftime("%Y%m%d%H%M%S", time.localtime())
    filenames = next(os.walk(folder_path), (None, None, []))[2]
    data = []
    output_file = f'{output_path}/{current_time}mixed-result.xlsx'
    for file in filenames:
        if 'result' in file:
            try:
                df = pd.read_excel(f'{folder_path}/{file}', sheet_name='result', engine='openpyxl')
                for i in range(len(df['Instance'])):
                    data.append([
                        df['Instance'][i], 
                        df['run_time'][i], 
                        df['cost'][i], 
                        df['work_time'][i], 
                        str(df['version'][i])+'|'+str(df['params'][i])+'|'+ str(df['number_of_tech'][i]),
                        df['t_route'][i], 
                        df['uav_route'][i]
                    ])
            except FileNotFoundError:
                print(f"No such file or directory: {output_file}")
    df1= pd.DataFrame(data, columns=['Instance', 'run_time', 'cost', 'work_time', 'params', 't_route', 'uav_route'])
    to_excel(output_file, 'result', df1)

def convert_result_to_col(filename, output_file, run_num):
    # function definition
    def highlight_cols(x):
        
        # copy df to new - original data is not changed
        df = x.copy()
        df.loc[:, :] = 'background-color: white'
        # cols = []
        # for i in range (1, df.shape[1], 8):
        #     cols.extend([x for x in range(i, i+4)])
        cols = list(itertools.chain.from_iterable(list(map(lambda x: [f'run_time{x}', f'cost{x}', f'work_time{x}', f'params{x}'], range(0, (df.shape[1]-1)//4, 2)))))
            
        # overwrite values grey color
        df[cols] = 'background-color: #D9D9D9'
        # print(cols)
        # return color df
        return df
    try:
        data = []
        param_set = []
        instance_set = {}
        col_set = ['Instance']
        df = pd.read_excel(filename, sheet_name='result', engine='openpyxl')
        for i in df['params']:
            ip = i.split('|')[:4] + [i.split('|')[4].split('/')[1]]
            if ip not in param_set:
                param_set.append(ip)
        for i in df['Instance']:
            if i not in instance_set:
                instance_set[i] = [[] for x in range(len(param_set))]
        for i in range(len(df['Instance'])):
            idp = param_set.index(df['params'][i].split('|')[:4] + [df['params'][i].split('|')[4].split('/')[1]])
            ins = df['Instance'][i]
            instance_set[ins][idp].append([df['run_time'][i], df['cost'][i], df['work_time'][i], df['params'][i]])
        missing_set=[]
        for i in instance_set:
            for j in range(run_num):
                row_data = [i]
                for k in range(len(param_set)):
                    if len(instance_set[i][k]) > j:
                        row_data.extend(instance_set[i][k][j])
                    else:
                        row_data.extend(['-']*4)
                        miss = [i, *param_set[k]]
                        if miss not in missing_set:
                            missing_set.append(miss)
                data.append(row_data)
        pprint.pprint(missing_set)
        # col_set.extend(['run_time', 'cost', 'work_time', 'params']*len(param_set))
        sub_col_set = list(map(lambda x: [f'run_time{x}', f'cost{x}', f'work_time{x}', f'params{x}'], range(len(param_set))))
        col_set.extend(list(itertools.chain.from_iterable(sub_col_set)))
        df1 = pd.DataFrame(data, columns=col_set)
        # highlight_cols(df1)
        styled = (df1.style.apply(highlight_cols, axis=None))
        styled.to_excel(output_file, engine='openpyxl')
        # to_excel(output_file, 'result', df1)
    except FileNotFoundError:
        print(f"No such file or directory: {output_file}")

def combine_stats(folder_path, output_path):
    current_time = time.strftime("%Y%m%d%H%M%S", time.localtime())
    filenames = next(os.walk(folder_path), (None, None, []))[2]
    upper_data = []
    diff_data = []
    output_file = f'{output_path}/{current_time}mixed-stats.xlsx'
    for file in filenames:
        if 'stats' in file:
            try:
                df = pd.read_excel(f'{folder_path}/{file}', sheet_name='upper', engine='openpyxl')
                old_data = df.values.tolist()
                upper_data.extend(list(map(lambda i: old_data[i][1:], range(len(old_data)))))
                df1 = pd.read_excel(f'{folder_path}/{file}', sheet_name='diff', engine='openpyxl')
                old_data = df1.values.tolist()
                diff_data.extend(list(map(lambda i: old_data[i][1:], range(len(old_data)))))
                col_num = df.columns.shape[0]-1
            except FileNotFoundError:
                print(f"No such file or directory: {output_file}")
    iter_cols = list(map(lambda x: f'iter{x}', range(col_num-6)))
    df2 = pd.DataFrame(upper_data, columns=['Instance', 'version', 'run_time', 'tech_num', 'work_time', 'params', *iter_cols])
    to_excel(output_file, 'upper', df2)
    df3 = pd.DataFrame(diff_data, columns=['Instance', 'version', 'run_time', 'tech_num', 'work_time', 'params', *iter_cols])
    to_excel(output_file, 'diff', df3)

def convert_stats_to_result(folder_path, output_path):
    current_time = time.strftime("%Y%m%d%H%M%S", time.localtime())
    data = []
    output_file = f'{output_path}/{current_time}mixed-result.xlsx'
    try:
        df = pd.read_excel(f'{folder_path}', sheet_name='upper', engine='openpyxl')
        old_data = df.values.tolist()
        data.extend(list(map(lambda i: [old_data[i][1], old_data[i][3], min(old_data[i][7:]), old_data[i][5], str(old_data[i][2])+'|'+str(old_data[i][6])+'|'+ str(old_data[i][4])], range(len(old_data)))))
        
    except FileNotFoundError:
        print(f"No such file or directory: {output_file}")
    df1= pd.DataFrame(data, columns=['Instance', 'run_time', 'cost', 'work_time', 'params'])
    to_excel(output_file, 'result', df1)

def get_converge_iteration(folder_path, output_path):
    current_time = time.strftime("%Y%m%d%H%M%S", time.localtime())
    filenames = next(os.walk(folder_path), (None, None, []))[2]
    upper_data = []
    output_file = f'{output_path}/{current_time}converge_iters.xlsx'
    for file in filenames:
        if 'stats' in file:
            try:
                df = pd.read_excel(f'{folder_path}/{file}', sheet_name='upper', engine='openpyxl')
                col_num = df.columns.shape[0]-1
                old_data = df.values.tolist()
                upper_data.extend(list(map(lambda i: old_data[i][1:7] + [min(old_data[i][7:col_num]), old_data[i][7:col_num].index(min(old_data[i][7:col_num]))], range(len(old_data)))))
            except FileNotFoundError:
                print(f"No such file or directory: {output_file}")
    # iter_cols = list(map(lambda x: f'iter{x}', range(col_num-6)))
    df2 = pd.DataFrame(upper_data, columns=['Instance', 'version', 'run_time', 'tech_num', 'work_time', 'params', 'cost', 'iter'])
    to_excel(output_file, 'upper', df2)

if __name__ == "__main__":
    # i1 = [2, 8, 5, 7, 3, 4, 6, 1]
    # i2 = [6, 3, 4, 2, 8, 7, 1, 5]
    # ds = hamming_distance(i1, i2, 6)
    # print(ds)
    # save_stats('instance', 'version', 'runtime', 'tech_num', 'work_time', 'upper', [0,1])
    # adjust_results('../result6-20.xlsx', '../mixed-result.xlsx')
    # convert_result_to_col('../result3/general/20211225082240mixed-result.xlsx', '../compare-result.xlsx', 5)
    # combine_results('../result3/new', '../result3/general')
    # combine_stats('../result3/old', '../result3/general')
    # convert_stats_to_result('../result3/general/20211211015719mixed-stats.xlsx', '../result3/general')
    # get_converge_iteration('../result3/general/21_12_17', '../result3/general/21_12_17')
    pass