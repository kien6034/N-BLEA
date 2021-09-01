from copy import Error
from matplotlib import markers
import matplotlib.pyplot as plt
import matplotlib as mpl
import time
import pandas as pd
from NBLEA import Low_GA


def to_excel(filename, data_ws, df, start_row=2, start_col=2):
    """Replacement for pandas .to_excel(). 

    For .xlsx and .xls formats only.

    args:
        start_row: df row +2; does not include header and is 1 based indexed.
    """
    writer = pd.ExcelWriter(filename.lower(), engine='openpyxl')
    import openpyxl
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    if data_ws not in wb.sheetnames:
        wb.create_sheet(data_ws)

    # Create the worksheet if it does not yet exist.
    writer.book = wb
    writer.sheets = {x.title: x for x in wb.worksheets}

    ws = writer.sheets[data_ws]
    # Fill with blanks.
    try:
        for row in ws:
            for cell in row:
                cell.value = None
    except TypeError:
        pass

    # Write manually to avoid overwriting formats.

    # Column names.
    ws.cell(1, 1).value = df.columns.name
    for icol, col_name in zip(range(2, len(df.columns) + 2), df.columns):
        ws.cell(1, icol).value = col_name

    # Row headers.
    for irow, row_name in zip(range(2, len(df.index) + 2), df.index):
        ws.cell(irow, 1).value = row_name

    # Body cells.
    for row, irow in zip([x[1] for x in df.iloc[start_row - 2:].iterrows()], list(range(start_row, len(df.index) + 2))):
        for cell, icol in zip([x[1] for x in row.iloc[start_col - 2:].items()], list(range(start_col, len(df.columns) + 2))):
            ws.cell(irow, icol).value = cell  # Skip the index.

    for row in ws.values:
        print('\t'.join(str(x or '') for x in row))
    print('Saving.')
    while True:
        try:
            writer.save()
            break
        except PermissionError:
            print(f'Please close {filename} before we can write to it!')
            time.sleep(2)
    writer.close()
    print('Done saving df.')

def get_result(graph, time_diff, best_fitness, best_t_tour, best_uav_tour, best_route_details):
    specific_route = Low_GA.get_specific_route(best_t_tour)
    sorted_route, t_back_time, max_cost = Low_GA.sort_by_time(graph, specific_route)
    last_node = list(sorted_route.keys())[len(sorted_route)-1]
    total_work_time = best_route_details['time_at_node'][last_node][0] + graph.t_time[last_node][0]
    active_tech = len(Low_GA.get_specific_route(best_t_tour))
    
    return graph.fileName, time_diff, active_tech, best_fitness, total_work_time, str(best_t_tour), str(best_uav_tour)

def save_solution(instance, run_time, number_of_tech, total_wait_time, total_work_time, tech_tour, uav_tour):
    # Instance, run_time, number_of_tech, total_wait_time, total_work_time, tech_tour, uav_tour
    inputFile = instance.replace('Instances/', '')
    inputFileName = inputFile.replace('.txt', '')
    print(inputFileName)
    data=[]
    try:
        xl = pd.ExcelFile('result.xlsx')
        if inputFileName in xl.sheet_names:
            df = pd.read_excel('result.xlsx', sheet_name=inputFileName)
            
            for i in range(len(df['Instance'])):
                data.append([
                    df['Instance'][i], df['run_time'][i], df['number_of_tech'][i], df['total_wait_time'][i], df['total_work_time'][i], df['tech_tour'][i], df['uav_tour'][i]
                ])
    except FileNotFoundError:
        print("No such file or directory: 'result.xlsx'")
    data.append([inputFileName, run_time, number_of_tech, total_wait_time, total_work_time, tech_tour, uav_tour])
    df = pd.DataFrame(data, columns=['Instance', 'run_time', 'number_of_tech', 'total_wait_time', 'total_work_time', 'tech_tour', 'uav_tour'])
    to_excel('result.xlsx', inputFileName, df)

def plot_solution():
    import numpy as np
    df = pd.read_excel('data.xlsx', sheet_name='GA')
    df1 = pd.read_excel('data.xlsx', sheet_name='MILP')

    data1 = []
    data2 = []
    data3 = []

    for i in range(len(df1['cost'])):
        data3.append(df1['cost'][i])

    tol = 0
    index_arr1 = df.loc[df['num_of_nodes'] == 10].index.values.tolist()
    min_cost = df['cost'][index_arr1[0]]
    for i in index_arr1:
        tol += df['cost'][i]
        min_cost = min(min_cost, df['cost'][i])
    data1.append(tol / len(df.loc[df['num_of_nodes'] == 10]))
    data2.append(min_cost)

    tol = 0
    index_arr2 = df.loc[df['num_of_nodes'] == 15].index.values.tolist()
    min_cost = df['cost'][index_arr2[0]]
    for i in index_arr2:
        tol += df['cost'][i]
        min_cost = min(min_cost, df['cost'][i])
    data1.append(tol / len(df.loc[df['num_of_nodes'] == 15]))
    data2.append(min_cost)

    tol = 0
    index_arr3 = df.loc[df['num_of_nodes'] == 20].index.values.tolist()
    min_cost = df['cost'][index_arr3[0]]
    for i in index_arr3:
        tol += df['cost'][i]
        min_cost = min(min_cost, df['cost'][i])
    data1.append(tol / len(df.loc[df['num_of_nodes'] == 20]))
    data2.append(min_cost)

    width = 0.3
    labels=['Customer = 10', 'Customer = 15', 'Customer = 20']
    plt.bar(np.arange(len(data1)), data1, width=width)
    plt.bar(np.arange(len(data2)) + width, data2, width=width)
    plt.bar(np.arange(len(data2)) + width * 2, data3, width=width)
    plt.legend(labels=['GA average cost', 'GA best cost', 'MILP'])
    plt.xticks([0 + width, 1 + width, 2 + width], labels)
    plt.ylabel('Cost')
    plt.show()

if __name__ == "__main__":
    plot_solution()
    pass